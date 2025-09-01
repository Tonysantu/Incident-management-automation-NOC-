from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import Dict
from datetime import datetime

LOG_FILE = Path("incident_log.xlsx")

def log_incident(alert: Dict, incident_id: str) -> None:
    headers = ["Incident ID", "Severity", "Message", "Source", "Created Time"]
    if LOG_FILE.exists():
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws.append([incident_id, alert.get("severity"), alert.get("message"),
               alert.get("source"), datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(LOG_FILE)
    print(f"[log] Logged incident {incident_id} to {LOG_FILE}")
